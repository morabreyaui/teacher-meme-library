"""Build the Teacher AI Tools ideas spreadsheet for the Monday pitch."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/morabreyaui/Code/teacher-meme-generator/Teacher AI Tools - Ideas.xlsx"

ROWS = [
    {
        "n": 1,
        "idea": '"What Kind of Teacher Are You?" Quiz',
        "status": "Built",
        "what": "Short personality quiz; teachers get a fun teaching-style result they can share.",
        "objective": "Acquisition (top of funnel). Highly shareable on social — drives non-users to a Legends-branded experience with zero friction.",
        "url": "http://localhost:3003",
        "url_note": "running",
    },
    {
        "n": 2,
        "idea": "General Meme Generator",
        "status": "Built",
        "what": "Classroom-life memes (no Legends branding required). Pick a situation, tone & character → generate.",
        "objective": "Acquisition + brand exposure. Casts the widest net: any teacher can use it, even non-users. They land on a Legends property and get exposed to the brand.",
        "url": "http://localhost:3001",
        "url_note": "running",
    },
    {
        "n": 3,
        "idea": "Branded Meme Generator",
        "status": "Built",
        "what": "Same flow as the general one, but starring Awakening avatars, beasties & classroom scenes.",
        "objective": "Branding + growth loop. Spreads our visual identity (characters, beasties) into teacher communities. Gives existing users something fun to share that only Legends can offer.",
        "url": "http://localhost:3002",
        "url_note": "running",
    },
    {
        "n": 4,
        "idea": "Custom Image-to-Meme",
        "status": "Planned (upgrade to #3)",
        "what": "User uploads a photo (their classroom, their actual avatar, their beastie) → AI inserts it into a meme situation.",
        "objective": 'Personalization → virality. Matt flagged this as "the killer part." People share what feels personal. Turns the tool from "cute" into "I have to send this to my coworker."',
        "url": "",
        "url_note": "",
    },
    {
        "n": 5,
        "idea": "Inspirational Quote / Poster Generator",
        "status": "New idea",
        "what": "Branded posters with motivational/teaching quotes + Legends visuals. Printable for the classroom.",
        "objective": "Brand presence in physical classrooms. Puts Legends visuals on the wall, every day, in front of every student. Long-tail brand impressions outside the screen.",
        "url": "",
        "url_note": "",
    },
    {
        "n": 6,
        "idea": "Educational Card Generator",
        "status": "New idea",
        "what": 'Branded "Did You Know?" cards or mini-lesson cards using beasties, tied to subjects/standards.',
        "objective": 'Retention + perceived value. Moves the suite from "fun toy" to "actually useful for teaching." Gives teachers a reason to come back during lesson prep, not just for laughs.',
        "url": "",
        "url_note": "",
    },
    {
        "n": 7,
        "idea": "Beastie Trading Cards (sweet spot — fun + rigor)",
        "status": "New idea",
        "what": 'Auto-generated Pokémon-style cards: student\'s beastie + a concept they mastered in Awakening (e.g., "Buzz the Bee — Pollination Master"). Stats tied to curriculum.',
        "objective": "Collectible loop tied to real learning. Combines fun (collectible, visual, social) with rigor (each card represents proven mastery). Closes the loop with the platform — the tool only works because the student is learning.",
        "url": "",
        "url_note": "",
    },
    {
        "n": 8,
        "idea": "Mastery Posters / Achievement Reels (sweet spot — fun + rigor)",
        "status": "New idea",
        "what": 'Auto-generated celebratory poster/reel triggered when a student masters a skill in Awakening (e.g., "I conquered fractions!" with their avatar).',
        "objective": "Product-led growth. Every learning event generates a shareable asset — parents post it, teachers send to families, students show classmates. The platform itself becomes the marketing engine.",
        "url": "",
        "url_note": "",
    },
]

HEADERS = ["#", "Idea", "Status", "What it is", "Objective / Why it matters", "Open (localhost)"]

# Style tokens
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # slate-800
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
ROW_FONT = Font(name="Calibri", size=11)
SWEET_SPOT_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber-100
BUILT_FILL = PatternFill("solid", fgColor="D1FAE5")  # emerald-100
PLANNED_FILL = PatternFill("solid", fgColor="DBEAFE")  # blue-100
NEW_FILL = PatternFill("solid", fgColor="F3E8FF")  # violet-100
LINK_FONT = Font(name="Calibri", size=11, color="2563EB", underline="single")
THIN = Side(border_style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
ws = wb.active
ws.title = "Ideas"

ws.append(HEADERS)
for col_idx, _ in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[1].height = 28

for r_offset, row in enumerate(ROWS, start=2):
    is_sweet = "sweet spot" in row["idea"]
    if is_sweet:
        row_fill = SWEET_SPOT_FILL
    elif row["status"] == "Built":
        row_fill = BUILT_FILL
    elif row["status"].startswith("Planned"):
        row_fill = PLANNED_FILL
    else:
        row_fill = NEW_FILL

    ws.cell(row=r_offset, column=1, value=row["n"])
    ws.cell(row=r_offset, column=2, value=row["idea"])
    ws.cell(row=r_offset, column=3, value=row["status"])
    ws.cell(row=r_offset, column=4, value=row["what"])
    ws.cell(row=r_offset, column=5, value=row["objective"])

    link_cell = ws.cell(row=r_offset, column=6)
    if row["url"]:
        display = row["url"]
        if row["url_note"]:
            display = f"{row['url']}\n({row['url_note']})"
        link_cell.value = display
        link_cell.hyperlink = row["url"]
        link_cell.font = LINK_FONT
    else:
        link_cell.value = "—"

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r_offset, column=col_idx)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = BORDER
        if col_idx != 6 or not row["url"]:
            cell.font = ROW_FONT
        cell.fill = row_fill

    if is_sweet:
        idea_cell = ws.cell(row=r_offset, column=2)
        idea_cell.font = Font(name="Calibri", size=11, bold=True)

widths = {1: 5, 2: 32, 3: 22, 4: 50, 5: 60, 6: 38}
for col_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

heights = {2: 60, 3: 60, 4: 70, 5: 75, 6: 65, 7: 75, 8: 95, 9: 85}
for r, h in heights.items():
    ws.row_dimensions[r].height = h

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(ROWS) + 1}"

# Legend sheet so the color coding makes sense at a glance
legend = wb.create_sheet("Legend")
legend.append(["Status", "Meaning"])
for col_idx in (1, 2):
    cell = legend.cell(row=1, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER

legend_rows = [
    ("Built", "App already exists and is running locally — link works.", BUILT_FILL),
    ("Planned", "Concrete next iteration of an existing app.", PLANNED_FILL),
    ("New idea", "Brainstormed concept; not started yet.", NEW_FILL),
    ("Sweet spot", "Fun + rigor crossover — strongest pitch for Vadim.", SWEET_SPOT_FILL),
]
for i, (status, meaning, fill) in enumerate(legend_rows, start=2):
    legend.cell(row=i, column=1, value=status).fill = fill
    legend.cell(row=i, column=2, value=meaning).fill = fill
    for col_idx in (1, 2):
        cell = legend.cell(row=i, column=col_idx)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
        cell.font = ROW_FONT

legend.column_dimensions["A"].width = 16
legend.column_dimensions["B"].width = 60

wb.save(OUTPUT_PATH)
print(f"Saved → {OUTPUT_PATH}")
